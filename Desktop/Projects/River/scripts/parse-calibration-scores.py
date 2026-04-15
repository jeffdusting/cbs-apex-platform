#!/usr/bin/env python3
"""
Parse Jeff's calibration scores.

Prefers docs/hyper-agent-v1/EVALUATOR_CALIBRATION_SCORING.xlsx (authoritative — Jeff's
submitted scoring sheet). Falls back to parsing EVALUATOR_CALIBRATION.md if the xlsx
is absent.

Outputs config/calibration-scores.json and prints a summary.
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

DIMENSIONS = [
    "KB Grounding",
    "Instruction Adherence",
    "Completeness",
    "Actionability",
    "Factual Discipline",
    "Risk Handling",
]

# Map of xlsx header text → canonical dimension name
XLSX_HEADER_TO_DIM = {
    "KB Grounding (25%)": "KB Grounding",
    "Instruction Adherence (20%)": "Instruction Adherence",
    "Completeness (15%)": "Completeness",
    "Actionability (15%)": "Actionability",
    "Factual Discipline (15%)": "Factual Discipline",
    "Risk Handling (10%)": "Risk Handling",
}

WEIGHTS = {
    "KB Grounding": 0.25,
    "Instruction Adherence": 0.20,
    "Completeness": 0.15,
    "Actionability": 0.15,
    "Factual Discipline": 0.15,
    "Risk Handling": 0.10,
}

PASS_THRESHOLD = 3.5

REPO_ROOT = Path(__file__).parent.parent
XLSX_PATH = REPO_ROOT / "docs" / "hyper-agent-v1" / "EVALUATOR_CALIBRATION_SCORING.xlsx"
MD_PATH = REPO_ROOT / "docs" / "hyper-agent-v1" / "EVALUATOR_CALIBRATION.md"
OUTPUT_PATH = REPO_ROOT / "config" / "calibration-scores.json"


def _normalise_score(val) -> int | None:
    """Accept int/float/str values and return an int in [1, 5] or None."""
    if val is None or val == "":
        return None
    try:
        n = int(round(float(val)))
    except (TypeError, ValueError):
        return None
    if 1 <= n <= 5:
        return n
    return None


def _weighted_composite(scores: dict[str, int]) -> float:
    """Re-normalise weights over dimensions actually scored."""
    if not scores:
        return 0.0
    total = sum(scores[d] * WEIGHTS[d] for d in scores)
    weight_sum = sum(WEIGHTS[d] for d in scores)
    return total / weight_sum if weight_sum else 0.0


def parse_xlsx(path: Path) -> list[dict]:
    """Extract scored outputs from Jeff's .xlsx scoring sheet."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("ERROR: openpyxl required to parse xlsx. pip install openpyxl.", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Calibration Scoring"] if "Calibration Scoring" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    col_index: dict[str, int] = {}
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        name = str(cell).strip()
        col_index[name] = idx

    # Required columns
    def col(name: str) -> int | None:
        return col_index.get(name)

    num_col = col("#")
    issue_col = col("Issue")
    role_col = col("Agent Role")
    task_col = col("Task Type")
    date_col = col("Date")
    notes_col = col("Notes")

    dim_cols: dict[str, int] = {}
    for header_text, dim in XLSX_HEADER_TO_DIM.items():
        idx = col(header_text)
        if idx is not None:
            dim_cols[dim] = idx

    outputs: list[dict] = []

    # Data rows — skip header (row 0). The description row has # = None, so it's filtered.
    for row in rows[1:]:
        if num_col is None or num_col >= len(row):
            continue
        num_val = row[num_col]
        if num_val is None:
            continue
        try:
            output_num = int(num_val)
        except (TypeError, ValueError):
            continue

        scores: dict[str, int] = {}
        for dim, idx in dim_cols.items():
            if idx < len(row):
                norm = _normalise_score(row[idx])
                if norm is not None:
                    scores[dim] = norm

        if not scores:
            continue

        composite = _weighted_composite(scores)

        def _cell(idx: int | None) -> str:
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        date_val = row[date_col] if date_col is not None and date_col < len(row) else None
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        elif date_val is None:
            date_str = ""
        else:
            date_str = str(date_val).strip()

        outputs.append({
            "output_number": output_num,
            "issue": _cell(issue_col),
            "agent_role": _cell(role_col),
            "task_type": _cell(task_col),
            "date": date_str,
            "scores": scores,
            "composite": round(composite, 2),
            "pass": composite >= PASS_THRESHOLD,
            "dimensions_scored": len(scores),
            "notes": _cell(notes_col),
        })

    outputs.sort(key=lambda o: o["output_number"])
    return outputs


def parse_md(path: Path) -> list[dict]:
    """Fallback: extract scores from EVALUATOR_CALIBRATION.md scoring tables."""
    text = path.read_text()
    outputs: list[dict] = []
    sections = re.split(r"^## Output (\d+)", text, flags=re.MULTILINE)
    for i in range(1, len(sections), 2):
        output_num = int(sections[i])
        content = sections[i + 1] if i + 1 < len(sections) else ""
        agent_role = _extract_field(content, "Agent role")
        task_type = _extract_field(content, "Task type")
        date = _extract_field(content, "Date")
        issue = _extract_field(content, "Issue")

        scores = {}
        for dim in DIMENSIONS:
            pattern = rf"\|\s*{re.escape(dim)}\s*\|\s*(\d)\s*\|"
            match = re.search(pattern, content)
            if match:
                scores[dim] = int(match.group(1))
        if not scores:
            continue

        composite = _weighted_composite(scores)
        outputs.append({
            "output_number": output_num,
            "issue": issue,
            "agent_role": agent_role,
            "task_type": task_type,
            "date": date,
            "scores": scores,
            "composite": round(composite, 2),
            "pass": composite >= PASS_THRESHOLD,
            "dimensions_scored": len(scores),
            "notes": "",
        })
    return outputs


def _extract_field(content: str, field: str) -> str:
    pattern = rf"\*\*{re.escape(field)}:\*\*\s*(.+)"
    match = re.search(pattern, content)
    if match:
        val = match.group(1).strip()
        if val.startswith("_") and val.endswith("_"):
            return ""
        return val
    return ""


def print_summary(outputs: list[dict]) -> None:
    scored = [o for o in outputs if o["dimensions_scored"] == len(DIMENSIONS)]
    partial = [o for o in outputs if 0 < o["dimensions_scored"] < len(DIMENSIONS)]

    print(f"\n{'='*60}")
    print(f"  EVALUATOR CALIBRATION SUMMARY")
    print(f"{'='*60}")
    print(f"  Fully scored outputs: {len(scored)} / 10")
    if partial:
        print(f"  Partially scored:     {len(partial)}")
    print()

    all_for_stats = outputs  # include partial for per-dim avg; compose uses normalisation

    if not outputs:
        print("  No scored outputs found.")
        print(f"{'='*60}\n")
        return

    print(f"  {'Dimension':<25} {'Avg':>5}  {'Min':>3}  {'Max':>3}  {'Weight':>6}")
    print(f"  {'-'*25} {'-'*5}  {'-'*3}  {'-'*3}  {'-'*6}")
    for dim in DIMENSIONS:
        vals = [o["scores"][dim] for o in all_for_stats if dim in o["scores"]]
        if vals:
            avg = sum(vals) / len(vals)
            print(f"  {dim:<25} {avg:5.2f}  {min(vals):3}  {max(vals):3}  {WEIGHTS[dim]:5.0%}")

    composites = [o["composite"] for o in outputs]
    avg_composite = sum(composites) / len(composites)
    passing = sum(1 for o in outputs if o["pass"])

    print()
    print(f"  Overall average composite: {avg_composite:.2f}")
    print(f"  Pass rate:                 {passing}/{len(outputs)} ({passing/len(outputs)*100:.0f}%)")
    print(f"  Pass threshold:            {PASS_THRESHOLD}")
    print(f"{'='*60}\n")

    print(f"  {'#':<3} {'Agent Role':<25} {'Composite':>9} {'Result':>7}  {'Dims'}")
    print(f"  {'-'*3} {'-'*25} {'-'*9} {'-'*7}  {'-'*4}")
    for o in outputs:
        result = "PASS" if o["pass"] else "FAIL"
        dims = f"{o['dimensions_scored']}/6"
        print(f"  {o['output_number']:<3} {o['agent_role']:<25} {o['composite']:9.2f} {result:>7}  {dims}")
    print()


def main():
    if XLSX_PATH.exists():
        source = XLSX_PATH
        outputs = parse_xlsx(XLSX_PATH)
        source_type = "xlsx"
    elif MD_PATH.exists():
        source = MD_PATH
        outputs = parse_md(MD_PATH)
        source_type = "markdown"
    else:
        print(f"Error: neither {XLSX_PATH} nor {MD_PATH} found", file=sys.stderr)
        sys.exit(1)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps({
        "source": str(source),
        "source_type": source_type,
        "parsed_at": datetime.now().isoformat(),
        "pass_threshold": PASS_THRESHOLD,
        "weights": WEIGHTS,
        "outputs": outputs,
    }, indent=2) + "\n")

    print(f"Wrote {len(outputs)} scored output(s) to {OUTPUT_PATH} (source: {source_type})")
    print_summary(outputs)


if __name__ == "__main__":
    main()
